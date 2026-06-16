"""
HR Candidate Follow-Up Automation
===================================
Reads candidate data from an Excel tracker, determines who needs a
follow-up based on status + days elapsed, and sends personalised emails
via Gmail SMTP. Also updates the tracker with sent timestamps.

Setup:
    pip install pandas openpyxl

Usage:
    python hr_followup_automation.py              # dry-run (no emails sent)
    python hr_followup_automation.py --send       # live email mode
    python hr_followup_automation.py --send --schedule  # daily auto-run

Environment variables (for live mode):
    HR_EMAIL      your Gmail address
    HR_EMAIL_PASS your Gmail App Password (not your login password)
"""

import os
import sys
import io
import smtplib
import logging
import argparse
import schedule
import time
import pandas as pd
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ---------------------------------------------------------------------------
# Fix Unicode encoding on Windows (cp1252 terminals)
# ---------------------------------------------------------------------------
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if sys.stderr.encoding and sys.stderr.encoding.lower() != "utf-8":
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
TRACKER_FILE = "candidate_tracker.xlsx"
SHEET_NAME   = "Candidates"
SMTP_HOST    = "smtp.gmail.com"
SMTP_PORT    = 587
HR_NAME      = "Disha"
ORG_NAME     = "JCF"

# How many days after last contact before a follow-up is triggered
FOLLOWUP_RULES = {
    "Applied":            2,
    "Shortlisted":        3,
    "Interview Scheduled": 1,
    "Under Review":       4,
}

# Statuses that should NEVER receive a follow-up
SKIP_STATUSES = {"Hired", "Rejected", "Withdrew", "On Hold"}

# ---------------------------------------------------------------------------
# Logging — UTF-8 safe file + console handlers
# ---------------------------------------------------------------------------
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

_fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

_console = logging.StreamHandler(sys.stdout)
_console.setFormatter(_fmt)

_file = logging.FileHandler("followup_log.txt", encoding="utf-8")
_file.setFormatter(_fmt)

logger.addHandler(_console)
logger.addHandler(_file)

# ---------------------------------------------------------------------------
# 1. READ CANDIDATE DATA
# ---------------------------------------------------------------------------
def load_candidates(filepath: str) -> pd.DataFrame:
    """Load candidate data from Excel tracker."""
    df = pd.read_excel(filepath, sheet_name=SHEET_NAME)
    df.columns = df.columns.str.strip()
    df["Last Contact Date"] = pd.to_datetime(df["Last Contact Date"], errors="coerce")
    df["Follow-Up Sent"]    = pd.to_datetime(df["Follow-Up Sent"],    errors="coerce")
    logger.info(f"Loaded {len(df)} candidates from {filepath}")
    return df

# ---------------------------------------------------------------------------
# 2. DETERMINE WHO NEEDS A FOLLOW-UP
# ---------------------------------------------------------------------------
def get_candidates_needing_followup(df: pd.DataFrame) -> pd.DataFrame:
    """Filter candidates who are overdue for a follow-up."""
    today    = datetime.today().date()
    eligible = []

    for _, row in df.iterrows():
        status = str(row.get("Status", "")).strip()

        # Skip terminal statuses
        if status in SKIP_STATUSES:
            continue

        # Skip if email is missing
        email = row.get("Email")
        if pd.isna(email) or str(email).strip() == "":
            logger.warning(f"  Skipping {row['Name']} -- no email address")
            continue

        # Skip if already followed up today
        last_sent = row.get("Follow-Up Sent")
        if pd.notna(last_sent) and last_sent.date() == today:
            continue

        # Determine required follow-up gap
        gap_days     = FOLLOWUP_RULES.get(status, 3)
        last_contact = row["Last Contact Date"]

        if pd.isna(last_contact):
            eligible.append(row)
            continue

        days_since = (today - last_contact.date()).days
        if days_since >= gap_days:
            eligible.append(row)

    result = pd.DataFrame(eligible)
    logger.info(f"{len(result)} candidate(s) need a follow-up today")
    return result

# ---------------------------------------------------------------------------
# 3. EMAIL TEMPLATES
# ---------------------------------------------------------------------------
EMAIL_TEMPLATES = {
    "Applied": (
        "Update on Your Application - {role} at {org}",
        """Dear {name},

Thank you for applying for the {role} position at {org}.
We have received your application and our team is currently reviewing it.
We aim to get back to you within the next 3-5 business days.

In the meantime, feel free to reach out if you have any questions.

Warm regards,
{hr_name}
HR Team, {org}
""",
    ),
    "Shortlisted": (
        "Great News! You've Been Shortlisted - {role} at {org}",
        """Dear {name},

We are pleased to inform you that you have been shortlisted for the
{role} role at {org}.

Our team will be in touch shortly to schedule the next steps.
Please keep an eye on your inbox.

Best regards,
{hr_name}
HR Team, {org}
""",
    ),
    "Interview Scheduled": (
        "Interview Reminder - {role} at {org}",
        """Dear {name},

This is a friendly reminder about your upcoming interview for the
{role} position at {org}.

Please confirm your availability by replying to this email.
If you need to reschedule, let us know at your earliest convenience.

Looking forward to speaking with you!

{hr_name}
HR Team, {org}
""",
    ),
    "Under Review": (
        "Application Status Update - {role} at {org}",
        """Dear {name},

We wanted to let you know that your application for {role} at {org}
is still under review. We appreciate your patience.

We expect to have an update for you soon and will be in touch directly.
Thank you for your continued interest in joining our team.

{hr_name}
HR Team, {org}
""",
    ),
    "default": (
        "Follow-Up: Your Application at {org}",
        """Dear {name},

We are following up regarding your application for the {role} position
at {org}.

Please reply to this email if you have any questions or updates on your end.

Best regards,
{hr_name}
HR Team, {org}
""",
    ),
}


def build_email(row: pd.Series) -> tuple:
    """Return (to_address, subject, html_body) for a candidate."""
    status       = str(row.get("Status", "default")).strip()
    template_key = status if status in EMAIL_TEMPLATES else "default"
    subject_tpl, body_tpl = EMAIL_TEMPLATES[template_key]

    ctx = {
        "name":    str(row.get("Name", "Candidate")).split()[0],
        "role":    str(row.get("Role Applied", "the open position")),
        "org":     ORG_NAME,
        "hr_name": HR_NAME,
    }

    subject = subject_tpl.format(**ctx)
    body    = body_tpl.format(**ctx)
    html    = f"<pre style='font-family:Arial,sans-serif;font-size:14px'>{body}</pre>"
    return str(row["Email"]).strip(), subject, html

# ---------------------------------------------------------------------------
# 4. SEND EMAIL
# ---------------------------------------------------------------------------
def send_email(
    to: str,
    subject: str,
    html_body: str,
    sender: str,
    password: str,
    dry_run: bool = True,
) -> bool:
    """Send a single HTML email. Returns True on success."""
    if dry_run:
        logger.info(f"  [DRY-RUN] Would email -> {to} | Subject: {subject}")
        return True

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = sender
    msg["To"]      = to
    msg.attach(MIMEText(html_body, "html"))

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.ehlo()
            server.starttls()
            server.login(sender, password)
            server.sendmail(sender, to, msg.as_string())
        logger.info(f"  [SENT] Email delivered -> {to}")
        return True
    except smtplib.SMTPAuthenticationError:
        logger.error(f"  [ERROR] Gmail authentication failed. Check HR_EMAIL and HR_EMAIL_PASS.")
        return False
    except Exception as e:
        logger.error(f"  [ERROR] Failed to send to {to}: {e}")
        return False

# ---------------------------------------------------------------------------
# 5. UPDATE TRACKER
# ---------------------------------------------------------------------------
def update_tracker(filepath: str, sent_names: list):
    """Stamp 'Follow-Up Sent' timestamp and update 'Last Contact Date' in Excel."""
    wb = load_workbook(filepath)
    ws = wb[SHEET_NAME]

    headers      = {cell.value: cell.column for cell in ws[1]}
    followup_col = headers.get("Follow-Up Sent")
    contact_col  = headers.get("Last Contact Date")
    name_col     = headers.get("Name")

    now_str   = datetime.now().strftime("%Y-%m-%d %H:%M")
    today_str = datetime.now().strftime("%Y-%m-%d")
    sent_fill = PatternFill("solid", start_color="D9F0D3")  # light green

    for row in ws.iter_rows(min_row=2):
        name_cell = row[name_col - 1] if name_col else None
        if name_cell and name_cell.value in sent_names:
            if followup_col:
                row[followup_col - 1].value = now_str
                row[followup_col - 1].fill  = sent_fill
            if contact_col:
                row[contact_col - 1].value = today_str

    wb.save(filepath)
    logger.info(f"Tracker updated for {len(sent_names)} candidate(s)")

# ---------------------------------------------------------------------------
# 6. MAIN PIPELINE
# ---------------------------------------------------------------------------
def run_pipeline(dry_run: bool = True):
    mode = "DRY RUN" if dry_run else "LIVE"
    logger.info(f"=== Follow-Up Pipeline Started | {mode} ===")

    sender   = os.getenv("HR_EMAIL",      "hr@jcf.org")
    password = os.getenv("HR_EMAIL_PASS", "")

    if not dry_run and not password:
        logger.error("HR_EMAIL_PASS environment variable is not set. Aborting live run.")
        return

    df          = load_candidates(TRACKER_FILE)
    to_followup = get_candidates_needing_followup(df)

    if to_followup.empty:
        logger.info("No follow-ups needed today.")
        return

    sent = []
    for _, row in to_followup.iterrows():
        to_addr, subject, body = build_email(row)
        ok = send_email(to_addr, subject, body, sender, password, dry_run=dry_run)
        if ok:
            sent.append(row["Name"])

    if sent and not dry_run:
        update_tracker(TRACKER_FILE, sent)

    logger.info(f"=== Pipeline Done: {len(sent)}/{len(to_followup)} follow-ups processed ===")

# ---------------------------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="HR Follow-Up Automation")
    parser.add_argument("--send",     action="store_true", help="Send real emails (default: dry-run)")
    parser.add_argument("--schedule", action="store_true", help="Run daily at 09:00 AM")
    args = parser.parse_args()

    dry_run = not args.send

    if args.schedule:
        schedule.every().day.at("09:00").do(run_pipeline, dry_run=dry_run)
        logger.info("Scheduler active -- running daily at 09:00 AM. Ctrl+C to stop.")
        while True:
            schedule.run_pending()
            time.sleep(60)
    else:
        run_pipeline(dry_run=dry_run)
